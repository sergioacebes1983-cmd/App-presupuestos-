import streamlit as st
from openpyxl import Workbook
from io import BytesIO

st.set_page_config(page_title="Presupuestos", layout="wide")

st.title("📊 Generador de Presupuestos con Fórmulas (1000 filas)")

# Función que crea Excel con fórmulas
def crear_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuestos"

    headers = [
        "Cliente",
        "Nº Vecinos",
        "Coste Tegui",
        "Nº Marcos",
        "Coste Marcos",
        "Nº Microaccess",
        "Coste Microaccess",
        "Nº Llaves",
        "Coste Llaves",
        "Coste Material",
        "Horas mano de obra (4 of.)",
        "Coste Personal",
        "Coste Total",
        "% Beneficio",
        "Beneficio Total",
        "Total Presupuesto",
        "Coste punto por vecino",
        "Número de presupuesto"
    ]
    ws.append(headers)

    # Insertamos 1000 filas con fórmulas
    for row in range(2, 1002):  # desde fila 2 hasta 1001
        ws.cell(row=row, column=5, value=f"=D{row}*120")   # Coste Marcos
        ws.cell(row=row, column=7, value=f"=F{row}*120")   # Coste Microaccess
        ws.cell(row=row, column=9, value=f"=H{row}*2.5")   # Coste Llaves
        ws.cell(row=row, column=10, value=f"=C{row}+E{row}+G{row}+I{row}")  # Coste Material
        ws.cell(row=row, column=12, value=f"=K{row}*100")  # Coste Personal
        ws.cell(row=row, column=13, value=f"=J{row}+L{row}")  # Coste Total
        ws.cell(row=row, column=15, value=f"=M{row}*O{row}/100")  # Beneficio Total
        ws.cell(row=row, column=16, value=f"=M{row}+N{row}")  # Total Presupuesto
        ws.cell(row=row, column=17, value=f"=P{row}/B{row}")  # Coste punto por vecino

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# Botón de descarga
st.download_button(
    "⬇️ Descargar Excel con fórmulas (1000 filas)",
    data=crear_excel(),
    file_name="Presupuesto_1000_con_formulas.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
